import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row  # this will return number of row in a sheet


def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column  ##this will return number of column in a sheet


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data #writing data in a excel sheet
    workbook.save(path)

