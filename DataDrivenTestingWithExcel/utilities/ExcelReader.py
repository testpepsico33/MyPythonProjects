import openpyxl


def get_data_from_excel(path, sheet_name):
    final_list=[]
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name] # sheet name of given excel
    total_row =sheet.max_row
    total_colum =sheet.max_column

    for r in range(2,total_row+1):
        row_list=[]
        for c in range(1,total_colum+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
    return final_list